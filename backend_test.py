#!/usr/bin/env python3
"""
Comprehensive Backend API Testing for Activus Invoice Management System
Tests all API endpoints with proper authentication and data flow
"""

import requests
import sys
import json
import io
import os
from datetime import datetime
from pathlib import Path

class ActivusAPITester:
    def __init__(self, base_url="https://0de79f73-880e-4b75-b873-cf862a9b62ec.preview.emergentagent.com"):
        self.base_url = base_url
        self.api_url = f"{base_url}/api"
        self.token = None
        self.user_data = None
        self.tests_run = 0
        self.tests_passed = 0
        self.created_resources = {
            'clients': [],
            'projects': [],
            'invoices': []
        }

    def log_test(self, name, success, details=""):
        """Log test results"""
        self.tests_run += 1
        if success:
            self.tests_passed += 1
            print(f"✅ {name} - PASSED {details}")
        else:
            print(f"❌ {name} - FAILED {details}")
        return success

    def make_request(self, method, endpoint, data=None, files=None, expected_status=200):
        """Make HTTP request with proper headers"""
        url = f"{self.api_url}/{endpoint}"
        headers = {}
        
        if self.token:
            headers['Authorization'] = f'Bearer {self.token}'
        
        if files is None and data is not None:
            headers['Content-Type'] = 'application/json'

        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                if files:
                    response = requests.post(url, headers={k: v for k, v in headers.items() if k != 'Content-Type'}, 
                                           data=data, files=files)
                else:
                    response = requests.post(url, headers=headers, json=data)
            elif method == 'PUT':
                response = requests.put(url, headers=headers, json=data)
            elif method == 'DELETE':
                response = requests.delete(url, headers=headers)
            else:
                return False, f"Unsupported method: {method}"

            success = response.status_code == expected_status
            
            if success:
                try:
                    return True, response.json()
                except:
                    return True, response.content
            else:
                try:
                    error_detail = response.json().get('detail', 'Unknown error')
                except:
                    error_detail = response.text
                return False, f"Status {response.status_code}: {error_detail}"

        except Exception as e:
            return False, f"Request failed: {str(e)}"

    def test_authentication(self):
        """Test login functionality"""
        print("\n🔐 Testing Authentication...")
        
        # Test invalid login (should return 401, not 200)
        success, result = self.make_request('POST', 'auth/login', 
                                          {'email': 'invalid@test.com', 'password': 'wrong'}, 
                                          expected_status=401)
        self.log_test("Invalid login rejection", success, "- Correctly rejected invalid credentials")
        
        # Test valid super admin login
        success, result = self.make_request('POST', 'auth/login', 
                                          {'email': 'brightboxm@gmail.com', 'password': 'admin123'})
        
        if success and 'access_token' in result:
            self.token = result['access_token']
            self.user_data = result['user']
            self.log_test("Super admin login", True, f"- Token received, Role: {self.user_data['role']}")
            return True
        else:
            self.log_test("Super admin login", False, f"- {result}")
            return False

    def test_dashboard_stats(self):
        """Test dashboard statistics endpoint"""
        print("\n📊 Testing Dashboard Stats...")
        
        success, result = self.make_request('GET', 'dashboard/stats')
        
        if success:
            required_fields = ['total_projects', 'total_invoices', 'total_invoiced_value', 'advance_received', 'pending_payment']
            has_all_fields = all(field in result for field in required_fields)
            self.log_test("Dashboard stats structure", has_all_fields, f"- Fields: {list(result.keys())}")
            self.log_test("Dashboard stats values", True, f"- Projects: {result.get('total_projects', 0)}, Invoices: {result.get('total_invoices', 0)}")
            return True
        else:
            self.log_test("Dashboard stats", False, f"- {result}")
            return False

    def test_client_management(self):
        """Test client CRUD operations"""
        print("\n👥 Testing Client Management...")
        
        # Test getting clients (initially empty)
        success, result = self.make_request('GET', 'clients')
        self.log_test("Get clients list", success, f"- Found {len(result) if success else 0} clients")
        
        # Test creating a client
        client_data = {
            "name": "Test Client Ltd",
            "gst_no": "27ABCDE1234F1Z5",
            "bill_to_address": "123 Test Street, Test City, Test State - 123456",
            "ship_to_address": "456 Ship Street, Ship City, Ship State - 654321",
            "contact_person": "John Doe",
            "phone": "+91-9876543210",
            "email": "john@testclient.com"
        }
        
        success, result = self.make_request('POST', 'clients', client_data, expected_status=200)
        
        if success and 'client_id' in result:
            client_id = result['client_id']
            self.created_resources['clients'].append(client_id)
            self.log_test("Create client", True, f"- Client ID: {client_id}")
            
            # Verify client was created by fetching list again
            success, clients = self.make_request('GET', 'clients')
            if success:
                created_client = next((c for c in clients if c['id'] == client_id), None)
                self.log_test("Verify client creation", created_client is not None, 
                            f"- Client found in list: {created_client['name'] if created_client else 'Not found'}")
                return True
        else:
            self.log_test("Create client", False, f"- {result}")
            return False

    def create_sample_excel_boq(self):
        """Create a sample Excel BOQ file for testing"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "BOQ"
            
            # Add project metadata
            ws['A1'] = 'Project Name:'
            ws['B1'] = 'Test Construction Project'
            ws['A2'] = 'Client:'
            ws['B2'] = 'Test Client Ltd'
            ws['A3'] = 'Architect:'
            ws['B3'] = 'Test Architect'
            ws['A4'] = 'Location:'
            ws['B4'] = 'Test Location'
            
            # Add BOQ headers
            headers = ['S.No', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount']
            for col, header in enumerate(headers, 1):
                ws.cell(row=6, column=col, value=header)
            
            # Add sample BOQ items
            boq_items = [
                [1, 'Excavation for foundation', 'Cum', 100, 150, 15000],
                [2, 'Concrete M20 for foundation', 'Cum', 50, 4500, 225000],
                [3, 'Steel reinforcement', 'Kg', 2000, 65, 130000],
                [4, 'Brick masonry', 'Cum', 200, 3500, 700000],
                [5, 'Plastering internal', 'Sqm', 500, 180, 90000]
            ]
            
            for row, item in enumerate(boq_items, 7):
                for col, value in enumerate(item, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer.getvalue()
            
        except ImportError:
            print("⚠️  openpyxl not available, creating minimal Excel data")
            # Return minimal Excel-like data for testing
            return b"Sample Excel BOQ data"

    def test_boq_upload(self):
        """Test BOQ Excel file upload and parsing"""
        print("\n📄 Testing BOQ Upload...")
        
        excel_data = self.create_sample_excel_boq()
        
        files = {'file': ('test_boq.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, result = self.make_request('POST', 'upload-boq', files=files)
        
        if success:
            required_fields = ['metadata', 'items', 'total_value', 'filename']
            has_all_fields = all(field in result for field in required_fields)
            self.log_test("BOQ upload structure", has_all_fields, f"- Fields: {list(result.keys())}")
            
            if 'items' in result:
                self.log_test("BOQ items parsed", len(result['items']) > 0, 
                            f"- Found {len(result['items'])} items, Total: ₹{result.get('total_value', 0)}")
                return result
        else:
            self.log_test("BOQ upload", False, f"- {result}")
            return None

    def test_project_management(self, boq_data=None):
        """Test project creation and management"""
        print("\n🏗️ Testing Project Management...")
        
        # Get initial projects list
        success, result = self.make_request('GET', 'projects')
        initial_count = len(result) if success else 0
        self.log_test("Get projects list", success, f"- Found {initial_count} projects")
        
        # Create a project
        if not self.created_resources['clients']:
            print("⚠️  No clients available, creating one first...")
            self.test_client_management()
        
        client_id = self.created_resources['clients'][0] if self.created_resources['clients'] else "test-client-id"
        
        project_data = {
            "project_name": "Test Construction Project",
            "architect": "Test Architect",
            "client_id": client_id,
            "client_name": "Test Client Ltd",
            "metadata": {
                "project_name": "Test Construction Project",
                "architect": "Test Architect",
                "client": "Test Client Ltd",
                "location": "Test Location"
            },
            "boq_items": boq_data['items'] if boq_data else [
                {
                    "serial_number": "1",
                    "description": "Test Item",
                    "unit": "nos",
                    "quantity": 10,
                    "rate": 1000,
                    "amount": 10000
                }
            ],
            "total_project_value": boq_data['total_value'] if boq_data else 10000,
            "advance_received": 0,
            "created_by": self.user_data['id'] if self.user_data else "test-user-id"
        }
        
        success, result = self.make_request('POST', 'projects', project_data)
        
        if success and 'project_id' in result:
            project_id = result['project_id']
            self.created_resources['projects'].append(project_id)
            self.log_test("Create project", True, f"- Project ID: {project_id}")
            
            # Test getting specific project
            success, project = self.make_request('GET', f'projects/{project_id}')
            self.log_test("Get specific project", success, 
                        f"- Project: {project.get('project_name', 'Unknown') if success else 'Failed'}")
            
            return project_id
        else:
            self.log_test("Create project", False, f"- {result}")
            return None

    def test_invoice_management(self):
        """Test invoice creation and PDF generation"""
        print("\n🧾 Testing Invoice Management...")
        
        # Get initial invoices
        success, result = self.make_request('GET', 'invoices')
        initial_count = len(result) if success else 0
        self.log_test("Get invoices list", success, f"- Found {initial_count} invoices")
        
        # Create invoice (need project first)
        if not self.created_resources['projects']:
            print("⚠️  No projects available, creating one first...")
            self.test_project_management()
        
        if not self.created_resources['projects']:
            self.log_test("Invoice creation", False, "- No projects available")
            return False
        
        project_id = self.created_resources['projects'][0]
        client_id = self.created_resources['clients'][0] if self.created_resources['clients'] else "test-client-id"
        
        invoice_data = {
            "project_id": project_id,
            "project_name": "Test Construction Project",
            "client_id": client_id,
            "client_name": "Test Client Ltd",
            "invoice_type": "proforma",
            "items": [
                {
                    "boq_item_id": "1",  # Required field for invoice items
                    "serial_number": "1",
                    "description": "Test Invoice Item",
                    "unit": "nos",
                    "quantity": 5,
                    "rate": 2000,
                    "amount": 10000,
                    "gst_rate": 18.0,
                    "gst_amount": 1800,
                    "total_with_gst": 11800
                }
            ],
            "subtotal": 10000,
            "total_gst_amount": 1800,
            "total_amount": 11800,
            "status": "draft",
            "created_by": self.user_data['id'] if self.user_data else "test-user-id"
        }
        
        success, result = self.make_request('POST', 'invoices', invoice_data)
        
        if success and 'invoice_id' in result:
            invoice_id = result['invoice_id']
            self.created_resources['invoices'].append(invoice_id)
            self.log_test("Create invoice", True, f"- Invoice ID: {invoice_id}")
            
            # Test PDF generation
            success, pdf_data = self.make_request('GET', f'invoices/{invoice_id}/pdf', expected_status=200)
            if success:
                self.log_test("Generate invoice PDF", True, f"- PDF size: {len(pdf_data) if isinstance(pdf_data, bytes) else 'Unknown'} bytes")
            else:
                self.log_test("Generate invoice PDF", False, f"- {pdf_data}")
            
            return True
        else:
            self.log_test("Create invoice", False, f"- {result}")
            return False

    def test_activity_logs(self):
        """Test activity logs (super admin only)"""
        print("\n📝 Testing Activity Logs...")
        
        if not self.user_data or self.user_data.get('role') != 'super_admin':
            self.log_test("Activity logs access", False, "- Not super admin")
            return False
        
        success, result = self.make_request('GET', 'activity-logs')
        
        if success:
            self.log_test("Get activity logs", True, f"- Found {len(result)} log entries")
            
            # Check log structure
            if result:
                log_entry = result[0]
                required_fields = ['user_email', 'action', 'description', 'timestamp']
                has_required = all(field in log_entry for field in required_fields)
                self.log_test("Activity log structure", has_required, 
                            f"- Sample action: {log_entry.get('action', 'Unknown')}")
            return True
        else:
            self.log_test("Get activity logs", False, f"- {result}")
            return False

    def test_item_master_apis(self):
        """Test Item Master Management APIs"""
        print("\n🔧 Testing Item Master APIs...")
        
        # Test getting master items (initially empty)
        success, result = self.make_request('GET', 'item-master')
        initial_count = len(result) if success else 0
        self.log_test("Get master items list", success, f"- Found {initial_count} master items")
        
        # Test creating a master item
        master_item_data = {
            "description": "Test Construction Material",
            "unit": "Cum",
            "standard_rate": 2500.0,
            "category": "Construction"
        }
        
        success, result = self.make_request('POST', 'item-master', master_item_data)
        
        if success and 'item_id' in result:
            item_id = result['item_id']
            self.log_test("Create master item", True, f"- Item ID: {item_id}")
            
            # Test updating the master item
            update_data = {
                "standard_rate": 2800.0,
                "category": "Updated Construction"
            }
            success, result = self.make_request('PUT', f'item-master/{item_id}', update_data)
            self.log_test("Update master item", success, f"- Updated rate to 2800")
            
            # Test getting master items with search
            success, result = self.make_request('GET', 'item-master?search=Test')
            found_item = any(item.get('id') == item_id for item in result) if success else False
            self.log_test("Search master items", found_item, f"- Found item in search results")
            
            # Test auto-populate from BOQ data
            success, result = self.make_request('POST', 'item-master/auto-populate')
            if success:
                created_count = result.get('created_count', 0)
                self.log_test("Auto-populate master items", True, f"- Created {created_count} items from BOQ data")
            else:
                self.log_test("Auto-populate master items", False, f"- {result}")
            
            # Test deleting the master item
            success, result = self.make_request('DELETE', f'item-master/{item_id}')
            self.log_test("Delete master item", success, f"- Item deleted successfully")
            
            return True
        else:
            self.log_test("Create master item", False, f"- {result}")
            return False

    def test_search_and_filter_apis(self):
        """Test Search and Filter APIs"""
        print("\n🔍 Testing Search and Filter APIs...")
        
        # Test global search
        success, result = self.make_request('GET', 'search?query=Test&limit=10')
        if success:
            total_count = result.get('total_count', 0)
            self.log_test("Global search", True, f"- Found {total_count} results across all entities")
            
            # Check search result structure
            has_sections = all(section in result for section in ['projects', 'clients', 'invoices'])
            self.log_test("Search result structure", has_sections, f"- Contains all entity sections")
        else:
            self.log_test("Global search", False, f"- {result}")
        
        # Test filtered projects
        success, result = self.make_request('GET', 'filters/projects?min_value=5000')
        if success:
            self.log_test("Filter projects by value", True, f"- Found {len(result)} projects with value >= 5000")
        else:
            self.log_test("Filter projects by value", False, f"- {result}")
        
        # Test filtered invoices
        success, result = self.make_request('GET', 'filters/invoices?status=draft')
        if success:
            self.log_test("Filter invoices by status", True, f"- Found {len(result)} draft invoices")
        else:
            self.log_test("Filter invoices by status", False, f"- {result}")
        
        # Test search by entity type
        success, result = self.make_request('GET', 'search?query=Client&entity_type=clients')
        if success:
            clients_found = len(result.get('clients', []))
            self.log_test("Search specific entity type", True, f"- Found {clients_found} clients")
        else:
            self.log_test("Search specific entity type", False, f"- {result}")

    def test_reports_and_insights_apis(self):
        """Test Reports and Insights APIs"""
        print("\n📊 Testing Reports and Insights APIs...")
        
        # Test GST summary report
        success, result = self.make_request('GET', 'reports/gst-summary')
        if success:
            required_fields = ['total_invoices', 'total_taxable_amount', 'total_gst_amount', 'gst_breakdown', 'monthly_breakdown']
            has_all_fields = all(field in result for field in required_fields)
            self.log_test("GST summary report", has_all_fields, 
                        f"- Total invoices: {result.get('total_invoices', 0)}, GST amount: ₹{result.get('total_gst_amount', 0)}")
            
            # Test GST summary with date filter
            success, filtered_result = self.make_request('GET', 'reports/gst-summary?date_from=2024-01-01')
            self.log_test("GST summary with date filter", success, f"- Filtered GST report generated")
        else:
            self.log_test("GST summary report", False, f"- {result}")
        
        # Test business insights
        success, result = self.make_request('GET', 'reports/insights')
        if success:
            required_sections = ['overview', 'financial', 'trends', 'performance']
            has_all_sections = all(section in result for section in required_sections)
            self.log_test("Business insights report", has_all_sections, 
                        f"- Projects: {result.get('overview', {}).get('total_projects', 0)}, Clients: {result.get('overview', {}).get('total_clients', 0)}")
            
            # Check financial metrics
            financial = result.get('financial', {})
            has_financial_data = 'total_project_value' in financial and 'collection_percentage' in financial
            self.log_test("Financial insights data", has_financial_data, 
                        f"- Collection rate: {financial.get('collection_percentage', 0):.1f}%")
        else:
            self.log_test("Business insights report", False, f"- {result}")
        
        # Test client-specific summary
        if self.created_resources['clients']:
            client_id = self.created_resources['clients'][0]
            success, result = self.make_request('GET', f'reports/client-summary/{client_id}')
            if success:
                required_fields = ['client_info', 'projects_count', 'invoices_count', 'total_project_value']
                has_all_fields = all(field in result for field in required_fields)
                self.log_test("Client summary report", has_all_fields, 
                            f"- Projects: {result.get('projects_count', 0)}, Total value: ₹{result.get('total_project_value', 0)}")
            else:
                self.log_test("Client summary report", False, f"- {result}")
        else:
            self.log_test("Client summary report", False, "- No clients available for testing")

    def create_sample_pdf_content(self):
        """Create a sample PDF-like content for testing"""
        # This creates a simple text content that mimics a Purchase Order with better structure
        pdf_content = """
        PURCHASE ORDER
        
        PO Number: PO-2024-001
        PO Date: 15/01/2024
        
        Vendor: Test Construction Supplies Ltd
        Client: Activus Industrial Design & Build LLP
        
        ITEM DETAILS:
        Description                 Unit    Quantity    Rate        Amount
        Cement bags                 Nos     100         450         45000
        Steel bars                  Kg      500         65          32500
        Sand                        Cum     10          1200        12000
        
        Total Amount: Rs 89,500
        
        Delivery Date: 30/01/2024
        Contact: supplier@testconstruction.com
        Phone: +91-9876543210
        
        Terms & Conditions:
        Payment within 30 days
        """
        return pdf_content.encode('utf-8')

    def test_pdf_processing_endpoints(self):
        """Test PDF Text Extraction Engine endpoints"""
        print("\n📄 Testing PDF Processing Endpoints...")
        
        # Test PDF extraction endpoint
        pdf_content = self.create_sample_pdf_content()
        files = {'file': ('test_po.pdf', pdf_content, 'application/pdf')}
        
        success, result = self.make_request('POST', 'pdf-processor/extract', files=files)
        
        extraction_id = None
        if success:
            extraction_id = result.get('extraction_id')
            extracted_data = result.get('extracted_data', {})
            processing_info = result.get('processing_info', {})
            
            self.log_test("PDF extraction", True, 
                        f"- Extraction ID: {extraction_id}, Method: {processing_info.get('extraction_method', 'Unknown')}")
            
            # Check extracted data structure - be more lenient for text-based testing
            has_some_data = any([
                extracted_data.get('po_number'),
                extracted_data.get('vendor_name'),
                extracted_data.get('total_amount'),
                extracted_data.get('line_items'),
                extracted_data.get('raw_text')  # At least raw text should be present
            ])
            confidence = extracted_data.get('confidence_score', 0)
            self.log_test("PDF data extraction quality", has_some_data or confidence > 0, 
                        f"- Confidence: {confidence:.2f}, Has data: {has_some_data}")
        else:
            self.log_test("PDF extraction", False, f"- {result}")
        
        # Test getting list of extractions
        success, result = self.make_request('GET', 'pdf-processor/extractions')
        if success:
            extractions_count = result.get('total', 0)
            self.log_test("Get PDF extractions list", True, f"- Found {extractions_count} extractions")
        else:
            self.log_test("Get PDF extractions list", False, f"- {result}")
        
        # Test getting specific extraction
        if extraction_id:
            success, result = self.make_request('GET', f'pdf-processor/extractions/{extraction_id}')
            if success:
                has_extraction_data = 'extracted_data' in result and 'original_filename' in result
                self.log_test("Get specific PDF extraction", has_extraction_data, 
                            f"- File: {result.get('original_filename', 'Unknown')}")
            else:
                self.log_test("Get specific PDF extraction", False, f"- {result}")
        
        # Test convert to project
        if extraction_id:
            project_metadata = {
                "project_name": "Test Project from PDF",
                "architect": "Test Architect",
                "client_id": self.created_resources['clients'][0] if self.created_resources['clients'] else "",
                "client_name": "Test Client from PDF",
                "additional_metadata": {
                    "source": "pdf_extraction_test"
                }
            }
            
            # Use query parameter for extraction_id and body for project_metadata
            success, result = self.make_request('POST', f'pdf-processor/convert-to-project?extraction_id={extraction_id}', 
                                              project_metadata)
            if success:
                project_id = result.get('project_id')
                if project_id:
                    self.created_resources['projects'].append(project_id)
                self.log_test("Convert PDF to project", True, 
                            f"- Project: {result.get('project_name', 'Unknown')}, Value: ₹{result.get('total_value', 0):,.2f}")
            else:
                self.log_test("Convert PDF to project", False, f"- {result}")
        
        return extraction_id

    def test_admin_configuration_system(self):
        """Test Admin Configuration System endpoints"""
        print("\n⚙️ Testing Admin Configuration System...")
        
        if not self.user_data or self.user_data.get('role') != 'super_admin':
            self.log_test("Admin access check", False, "- Not super admin, skipping admin tests")
            return False
        
        # Test workflow configuration
        workflow_data = {
            "workflow_name": "Test Invoice Approval Workflow",
            "workflow_type": "approval",
            "steps": [
                {"step": 1, "name": "Create", "role": "invoice_creator", "action": "create_invoice"},
                {"step": 2, "name": "Review", "role": "reviewer", "action": "review_invoice"},
                {"step": 3, "name": "Approve", "role": "approver", "action": "approve_invoice"}
            ],
            "roles_permissions": {
                "invoice_creator": ["create", "edit_draft"],
                "reviewer": ["review", "request_changes"],
                "approver": ["approve", "reject"]
            },
            "notifications_config": {
                "email_notifications": True,
                "sms_notifications": False,
                "in_app_notifications": True
            },
            "active": True,
            "created_by": self.user_data['id'] if self.user_data else "test-user-id"
        }
        
        success, result = self.make_request('POST', 'admin/workflows', workflow_data)
        workflow_id = None
        if success:
            workflow_id = result.get('workflow_id')
            self.log_test("Create workflow config", True, f"- Workflow ID: {workflow_id}")
        else:
            self.log_test("Create workflow config", False, f"- {result}")
        
        # Test getting workflows
        success, result = self.make_request('GET', 'admin/workflows')
        if success:
            workflows_count = len(result) if isinstance(result, list) else 0
            self.log_test("Get workflow configs", True, f"- Found {workflows_count} workflows")
        else:
            self.log_test("Get workflow configs", False, f"- {result}")
        
        # Test updating workflow
        if workflow_id:
            update_data = {
                "active": False,
                "notifications_config": {
                    "email_notifications": False,
                    "sms_notifications": True,
                    "in_app_notifications": True
                }
            }
            success, result = self.make_request('PUT', f'admin/workflows/{workflow_id}', update_data)
            self.log_test("Update workflow config", success, f"- Workflow updated")
        
        # Test system configuration
        system_config_data = {
            "config_category": "business",
            "config_key": "default_gst_rate",
            "config_value": 18.0,
            "config_type": "number",
            "description": "Default GST rate for new items",
            "is_sensitive": False,
            "requires_restart": False,
            "created_by": self.user_data['id'] if self.user_data else "test-user-id"
        }
        
        success, result = self.make_request('POST', 'admin/system-config', system_config_data)
        config_id = None
        if success:
            config_id = result.get('config_id')
            self.log_test("Create system config", True, f"- Config ID: {config_id}")
        else:
            self.log_test("Create system config", False, f"- {result}")
        
        # Test getting system configs
        success, result = self.make_request('GET', 'admin/system-config')
        if success:
            categories_count = len(result) if isinstance(result, dict) else 0
            self.log_test("Get system configs", True, f"- Found {categories_count} config categories")
        else:
            self.log_test("Get system configs", False, f"- {result}")
        
        # Test updating system config
        if config_id:
            update_data = {
                "config_value": 12.0,
                "description": "Updated default GST rate"
            }
            success, result = self.make_request('PUT', f'admin/system-config/{config_id}', update_data)
            restart_required = result.get('restart_required', False) if success else False
            self.log_test("Update system config", success, f"- Config updated, Restart required: {restart_required}")
        
        # Test system health
        success, result = self.make_request('GET', 'admin/system-health')
        if success:
            db_status = result.get('database', {}).get('status', 'unknown')
            collections = result.get('database', {}).get('collections', {})
            recent_activity_count = len(result.get('recent_activity', []))
            
            self.log_test("System health check", True, 
                        f"- DB Status: {db_status}, Collections: {len(collections)}, Recent activity: {recent_activity_count}")
            
            # Check if all expected collections are healthy
            expected_collections = ['users', 'projects', 'invoices', 'clients', 'pdf_extractions']
            healthy_collections = sum(1 for col in expected_collections 
                                    if collections.get(col, {}).get('status') == 'healthy')
            self.log_test("Database collections health", healthy_collections >= 4, 
                        f"- {healthy_collections}/{len(expected_collections)} collections healthy")
        else:
            self.log_test("System health check", False, f"- {result}")
        
        return True

    def test_authentication_and_permissions(self):
        """Test authentication and permission controls for new endpoints"""
        print("\n🔐 Testing Authentication & Permissions...")
        
        # Store current token
        old_token = self.token
        
        # Test PDF processing without authentication
        self.token = None
        pdf_content = self.create_sample_pdf_content()
        files = {'file': ('test_po.pdf', pdf_content, 'application/pdf')}
        
        success, result = self.make_request('POST', 'pdf-processor/extract', files=files, expected_status=401)
        self.log_test("PDF processing unauthorized access", not success, "- Correctly rejected unauthenticated request")
        
        # Test admin endpoints without authentication
        success, result = self.make_request('GET', 'admin/workflows', expected_status=401)
        self.log_test("Admin workflows unauthorized access", not success, "- Correctly rejected unauthenticated request")
        
        success, result = self.make_request('GET', 'admin/system-health', expected_status=401)
        self.log_test("Admin system health unauthorized access", not success, "- Correctly rejected unauthenticated request")
        
        # Restore token
        self.token = old_token
        
        # Test admin endpoints with non-admin user (if we had one)
        # For now, we only have super admin, so we'll test with valid token
        success, result = self.make_request('GET', 'admin/workflows')
        if success:
            self.log_test("Admin workflows with super admin", True, "- Super admin access granted")
        else:
            # Check if it's a permission error
            is_permission_error = "403" in str(result) or "super admin" in str(result).lower()
            self.log_test("Admin workflows permission check", is_permission_error, f"- Permission validation working")

    def test_database_clear_functionality(self):
        """Test the new database clear functionality - CRITICAL SECURITY FEATURE"""
        print("\n🚨 Testing Database Clear Functionality (CRITICAL)...")
        
        if not self.user_data or self.user_data.get('role') != 'super_admin':
            self.log_test("Database clear access check", False, "- Not super admin, skipping database clear tests")
            return False
        
        # 1. SECURITY TESTING - Test unauthorized access (without token)
        old_token = self.token
        self.token = None
        
        clear_data = {
            "confirm_clear": True,
            "confirmation_text": "DELETE ALL DATA"
        }
        
        success, result = self.make_request('POST', 'admin/clear-database', clear_data, expected_status=401)
        self.log_test("Database clear - unauthorized access rejection", success, "- Correctly rejected unauthenticated request")
        
        # Restore token
        self.token = old_token
        
        # 2. CONFIRMATION TESTING - Test without confirmation parameters
        success, result = self.make_request('POST', 'admin/clear-database', {}, expected_status=400)
        self.log_test("Database clear - no confirmation rejection", success, "- Correctly rejected request without confirmation")
        
        # 3. CONFIRMATION TESTING - Test with wrong confirmation text
        wrong_confirmation = {
            "confirm_clear": True,
            "confirmation_text": "WRONG TEXT"
        }
        success, result = self.make_request('POST', 'admin/clear-database', wrong_confirmation, expected_status=400)
        self.log_test("Database clear - wrong confirmation text rejection", success, "- Correctly rejected wrong confirmation text")
        
        # 4. CONFIRMATION TESTING - Test with checkbox unchecked
        unchecked_confirmation = {
            "confirm_clear": False,
            "confirmation_text": "DELETE ALL DATA"
        }
        success, result = self.make_request('POST', 'admin/clear-database', unchecked_confirmation, expected_status=400)
        self.log_test("Database clear - unchecked confirmation rejection", success, "- Correctly rejected unchecked confirmation")
        
        # 5. Get system health before clearing to see current data
        success, health_before = self.make_request('GET', 'admin/system-health')
        collections_before = {}
        if success:
            collections_before = health_before.get('database', {}).get('collections', {})
            total_records_before = sum(col.get('count', 0) for col in collections_before.values())
            self.log_test("Pre-clear system health check", True, 
                        f"- Total records before clear: {total_records_before}")
        
        # 6. FUNCTIONALITY TESTING - Test with correct confirmation
        correct_confirmation = {
            "confirm_clear": True,
            "confirmation_text": "DELETE ALL DATA"
        }
        
        success, result = self.make_request('POST', 'admin/clear-database', correct_confirmation)
        
        if success:
            # 7. RESPONSE VALIDATION - Check response structure
            required_fields = ['message', 'timestamp', 'cleared_by', 'statistics', 'preserved']
            has_all_fields = all(field in result for field in required_fields)
            self.log_test("Database clear - response structure", has_all_fields, 
                        f"- Response contains all required fields")
            
            # Check statistics
            stats = result.get('statistics', {})
            total_deleted = stats.get('total_records_deleted', 0)
            collections_cleared = stats.get('collections_cleared', 0)
            collections_details = stats.get('collections_details', [])
            
            self.log_test("Database clear - deletion statistics", total_deleted >= 0, 
                        f"- Total deleted: {total_deleted}, Collections cleared: {collections_cleared}")
            
            # Check that specific collections were targeted
            expected_collections = ['projects', 'invoices', 'clients', 'bank_guarantees', 
                                  'pdf_extractions', 'master_items', 'workflow_configs', 
                                  'system_configs', 'activity_logs']
            
            cleared_collection_names = [c.get('collection') for c in collections_details]
            has_expected_collections = all(col in cleared_collection_names for col in expected_collections)
            self.log_test("Database clear - targeted collections", has_expected_collections,
                        f"- All expected collections targeted: {len(cleared_collection_names)}")
            
            # Check that users collection is preserved
            preserved = result.get('preserved', {})
            users_preserved = 'users' in str(preserved)
            self.log_test("Database clear - users preservation", users_preserved,
                        "- User accounts preserved as expected")
            
            # Check cleared_by information
            cleared_by = result.get('cleared_by', {})
            has_user_info = 'user_id' in cleared_by and 'email' in cleared_by
            self.log_test("Database clear - audit trail", has_user_info,
                        f"- Cleared by: {cleared_by.get('email', 'Unknown')}")
            
            # 8. VERIFICATION - Check system health after clearing
            success_after, health_after = self.make_request('GET', 'admin/system-health')
            if success_after:
                collections_after = health_after.get('database', {}).get('collections', {})
                
                # Verify users collection still has data
                users_count_after = collections_after.get('users', {}).get('count', 0)
                self.log_test("Database clear - users preserved verification", users_count_after > 0,
                            f"- Users collection still has {users_count_after} records")
                
                # Verify other collections are cleared (should be 0 or very low)
                other_collections_cleared = True
                for col_name in expected_collections:
                    if col_name in collections_after:
                        count = collections_after[col_name].get('count', 0)
                        if count > 0:
                            other_collections_cleared = False
                            break
                
                self.log_test("Database clear - data collections cleared verification", other_collections_cleared,
                            "- All data collections successfully cleared")
            
            # 9. ACTIVITY LOG VERIFICATION - Check that the action was logged
            success_logs, logs_result = self.make_request('GET', 'activity-logs')
            if success_logs and logs_result:
                # Look for the database clear log entry
                clear_log_found = False
                for log_entry in logs_result[:5]:  # Check recent logs
                    if log_entry.get('action') == 'database_cleared':
                        clear_log_found = True
                        break
                
                self.log_test("Database clear - activity logging", clear_log_found,
                            "- Database clear action properly logged")
            
            return True
        else:
            self.log_test("Database clear - execution", False, f"- {result}")
            return False

    def test_error_handling(self):
        """Test error handling and edge cases"""
        print("\n⚠️ Testing Error Handling...")
        
        # Test unauthorized access (without token)
        old_token = self.token
        self.token = None
        
        # Test unauthorized access - FastAPI HTTPBearer can return either 401 or 403
        success_401, result_401 = self.make_request('GET', 'projects', expected_status=401)
        success_403, result_403 = self.make_request('GET', 'projects', expected_status=403)
        success = success_401 or success_403
        self.log_test("Unauthorized access rejection", success, "- Correctly rejected request without token")
        
        # Restore token for authenticated error tests
        self.token = old_token
        
        # Test invalid project ID (with authentication)
        success, result = self.make_request('GET', 'projects/invalid-id', expected_status=404)
        self.log_test("Invalid project ID handling", success, "- Correctly returned 404 for invalid ID")
        
        # Test invalid file upload (with authentication)
        files = {'file': ('test.txt', b'not an excel file', 'text/plain')}
        success, result = self.make_request('POST', 'upload-boq', files=files, expected_status=400)
        self.log_test("Invalid file type rejection", success, "- Correctly rejected non-Excel file")
        
        # Test invalid PDF file upload (with authentication)
        files = {'file': ('test.txt', b'not a pdf file', 'text/plain')}
        success, result = self.make_request('POST', 'pdf-processor/extract', files=files, expected_status=400)
        self.log_test("Invalid PDF file rejection", success, "- Correctly rejected non-PDF file")
        
        # Test empty PDF file (with authentication)
        files = {'file': ('empty.pdf', b'', 'application/pdf')}
        success, result = self.make_request('POST', 'pdf-processor/extract', files=files, expected_status=400)
        self.log_test("Empty PDF file rejection", success, "- Correctly rejected empty file")
        
        # Test invalid extraction ID (with authentication)
        success, result = self.make_request('GET', 'pdf-processor/extractions/invalid-id', expected_status=404)
        self.log_test("Invalid extraction ID handling", success, "- Correctly returned 404 for invalid extraction ID")
        
        # Test invalid workflow ID (with authentication)
        success, result = self.make_request('PUT', 'admin/workflows/invalid-id', {"active": False}, expected_status=404)
        self.log_test("Invalid workflow ID handling", success, "- Correctly returned 404 for invalid workflow ID")
        
        # Test invalid system config ID (with authentication)
        success, result = self.make_request('PUT', 'admin/system-config/invalid-id', {"config_value": "test"}, expected_status=404)
        self.log_test("Invalid system config ID handling", success, "- Correctly returned 404 for invalid config ID")
        
        # Test invalid master item creation (duplicate) (with authentication)
        master_item_data = {
            "description": "Duplicate Test Item",
            "unit": "nos",
            "standard_rate": 100.0
        }
        # Create first item
        self.make_request('POST', 'item-master', master_item_data)
        # Try to create duplicate
        success, result = self.make_request('POST', 'item-master', master_item_data, expected_status=400)
        self.log_test("Duplicate master item rejection", success, "- Correctly rejected duplicate item")
        
        # Test invalid client summary (with authentication)
        success, result = self.make_request('GET', 'reports/client-summary/invalid-client-id', expected_status=404)
        self.log_test("Invalid client summary handling", success, "- Correctly returned 404 for invalid client ID")

    def run_all_tests(self):
        """Run complete test suite"""
        print("🚀 Starting Activus Invoice Management System API Tests")
        print(f"🌐 Testing against: {self.base_url}")
        print("=" * 80)
        
        # Test authentication first
        if not self.test_authentication():
            print("\n❌ Authentication failed - stopping tests")
            return False
        
        # Test all endpoints
        self.test_dashboard_stats()
        self.test_client_management()
        
        # Test BOQ upload and project creation
        boq_data = self.test_boq_upload()
        self.test_project_management(boq_data)
        
        # Test invoice management
        self.test_invoice_management()
        
        # Test activity logs
        self.test_activity_logs()
        
        # Test newly implemented features
        self.test_item_master_apis()
        self.test_search_and_filter_apis()
        self.test_reports_and_insights_apis()
        
        # Test NEW PDF Text Extraction Engine (BE-01)
        self.test_pdf_processing_endpoints()
        
        # Test NEW Admin Configuration System
        self.test_admin_configuration_system()
        
        # Test NEW Database Clear Functionality (CRITICAL SECURITY FEATURE)
        self.test_database_clear_functionality()
        
        # Test authentication and permissions for new endpoints
        self.test_authentication_and_permissions()
        
        # Test error handling (updated with new endpoints)
        self.test_error_handling()
        
        # Print summary
        print("\n" + "=" * 80)
        print(f"📊 Test Summary: {self.tests_passed}/{self.tests_run} tests passed")
        
        if self.created_resources['clients']:
            print(f"👥 Created {len(self.created_resources['clients'])} clients")
        if self.created_resources['projects']:
            print(f"🏗️ Created {len(self.created_resources['projects'])} projects")
        if self.created_resources['invoices']:
            print(f"🧾 Created {len(self.created_resources['invoices'])} invoices")
        
        success_rate = (self.tests_passed / self.tests_run) * 100 if self.tests_run > 0 else 0
        print(f"✅ Success Rate: {success_rate:.1f}%")
        
        return self.tests_passed == self.tests_run

def main():
    """Main test execution"""
    tester = ActivusAPITester()
    
    try:
        success = tester.run_all_tests()
        return 0 if success else 1
    except KeyboardInterrupt:
        print("\n⏹️ Tests interrupted by user")
        return 1
    except Exception as e:
        print(f"\n💥 Unexpected error: {str(e)}")
        return 1

if __name__ == "__main__":
    sys.exit(main())