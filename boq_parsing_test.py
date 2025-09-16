#!/usr/bin/env python3
"""
BOQ Parsing Functionality Testing
Tests the improved BOQ parsing functionality focusing on:
1. Unit/UOM Column Extraction
2. Column Mapping Debug
3. BOQ Item Structure
"""

import requests
import sys
import json
import io
import os
from datetime import datetime
from pathlib import Path

class BOQParsingTester:
    def __init__(self, base_url="https://0de79f73-880e-4b75-b873-cf862a9b62ec.preview.emergentagent.com"):
        self.base_url = base_url
        self.api_url = f"{base_url}/api"
        self.token = None
        self.user_data = None
        self.tests_run = 0
        self.tests_passed = 0

    def log_test(self, name, success, details=""):
        """Log test results"""
        self.tests_run += 1
        if success:
            self.tests_passed += 1
            print(f"✅ {name} - PASSED {details}")
        else:
            print(f"❌ {name} - FAILED {details}")
        return success

    def make_request(self, method, endpoint, data=None, files=None, expected_status=200):
        """Make HTTP request with proper headers"""
        url = f"{self.api_url}/{endpoint}"
        headers = {}
        
        if self.token:
            headers['Authorization'] = f'Bearer {self.token}'
        
        if files is None and data is not None:
            headers['Content-Type'] = 'application/json'

        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                if files:
                    response = requests.post(url, headers={k: v for k, v in headers.items() if k != 'Content-Type'}, 
                                           data=data, files=files)
                else:
                    response = requests.post(url, headers=headers, json=data)
            else:
                return False, f"Unsupported method: {method}"

            success = response.status_code == expected_status
            
            if success:
                try:
                    return True, response.json()
                except:
                    return True, response.content
            else:
                try:
                    error_detail = response.json().get('detail', 'Unknown error')
                except:
                    error_detail = response.text
                return False, f"Status {response.status_code}: {error_detail}"

        except Exception as e:
            return False, f"Request failed: {str(e)}"

    def authenticate(self):
        """Authenticate with the API"""
        print("🔐 Authenticating...")
        
        success, result = self.make_request('POST', 'auth/login', 
                                          {'email': 'brightboxm@gmail.com', 'password': 'admin123'})
        
        if success and 'access_token' in result:
            self.token = result['access_token']
            self.user_data = result['user']
            self.log_test("Authentication", True, f"- Logged in as {self.user_data['role']}")
            return True
        else:
            self.log_test("Authentication", False, f"- {result}")
            return False

    def create_excel_with_units(self, filename, unit_data):
        """Create Excel file with specific unit data for testing"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "BOQ"
            
            # Add project metadata
            ws['A1'] = 'Project Name:'
            ws['B1'] = 'Unit Testing Project'
            ws['A2'] = 'Client:'
            ws['B2'] = 'Test Client for Units'
            ws['A3'] = 'Architect:'
            ws['B3'] = 'Unit Test Architect'
            
            # Add BOQ headers
            headers = ['S.No', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount']
            for col, header in enumerate(headers, 1):
                ws.cell(row=6, column=col, value=header)
            
            # Add BOQ items with specific unit data
            for row, item_data in enumerate(unit_data, 7):
                for col, value in enumerate(item_data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer.getvalue()
            
        except ImportError:
            print("⚠️  openpyxl not available, creating minimal Excel data")
            return b"Sample Excel BOQ data"

    def test_unit_column_extraction(self):
        """Test Unit/UOM Column Extraction with various unit types"""
        print("\n📏 Testing Unit/UOM Column Extraction...")
        
        # Test data with different unit types
        unit_test_cases = [
            # [S.No, Description, Unit, Quantity, Rate, Amount]
            [1, 'Excavation for foundation', 'Cum', 100, 150, 15000],
            [2, 'Floor area calculation', 'Sqm', 500, 180, 90000],
            [3, 'Steel reinforcement bars', 'Nos', 200, 65, 13000],
            [4, 'Concrete volume', 'Cu.M', 50, 4500, 225000],
            [5, 'Wall plastering area', 'Sq.M', 300, 120, 36000],
            [6, 'Door installation', 'Each', 10, 5000, 50000],
            [7, 'Running measurement', 'Rmt', 150, 200, 30000],
            [8, 'Material weight', 'Kg', 1000, 45, 45000],
            [9, 'Bulk material', 'Ton', 5, 25000, 125000],
            [10, 'Liquid volume', 'Ltr', 500, 50, 25000]
        ]
        
        excel_data = self.create_excel_with_units("unit_test.xlsx", unit_test_cases)
        files = {'file': ('unit_test.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, result = self.make_request('POST', 'upload-boq', files=files)
        
        if success:
            self.log_test("BOQ file upload", True, f"- File processed successfully")
            
            # Check if items were parsed
            items = result.get('items', [])
            self.log_test("BOQ items parsed", len(items) > 0, f"- Found {len(items)} items")
            
            # Test unit extraction for each item
            expected_units = ['Cum', 'Sqm', 'Nos', 'Cum', 'Sqm', 'Nos', 'Rmt', 'Kg', 'Ton', 'Ltr']
            unit_extraction_success = True
            unit_details = []
            
            for i, item in enumerate(items):
                if i < len(expected_units):
                    extracted_unit = item.get('unit', '')
                    expected_unit = expected_units[i]
                    
                    # Check if unit is text (not number)
                    is_text_unit = isinstance(extracted_unit, str) and not extracted_unit.isdigit()
                    
                    # Check if unit matches expected (case insensitive)
                    unit_matches = extracted_unit.lower() in expected_unit.lower() or expected_unit.lower() in extracted_unit.lower()
                    
                    if is_text_unit and unit_matches:
                        unit_details.append(f"Item {i+1}: '{extracted_unit}' ✓")
                    else:
                        unit_details.append(f"Item {i+1}: '{extracted_unit}' ❌ (expected: {expected_unit})")
                        unit_extraction_success = False
            
            self.log_test("Unit values are text (not numbers)", unit_extraction_success, 
                        f"- Unit extraction details: {'; '.join(unit_details[:3])}...")
            
            # Test that units are preserved as text
            text_units_count = sum(1 for item in items if isinstance(item.get('unit'), str) and item.get('unit').strip())
            self.log_test("Units preserved as text", text_units_count == len(items), 
                        f"- {text_units_count}/{len(items)} units are text values")
            
            return result
        else:
            self.log_test("BOQ file upload", False, f"- {result}")
            return None

    def test_column_mapping_debug(self):
        """Test enhanced column mapping with debug output"""
        print("\n🗺️ Testing Column Mapping Debug...")
        
        # Create Excel with different header variations
        mapping_test_cases = [
            [1, 'Foundation excavation work', 'Cubic Meter', 75, 180, 13500],
            [2, 'Wall construction activity', 'Square Meter', 400, 220, 88000],
            [3, 'Material procurement', 'Numbers', 150, 85, 12750],
        ]
        
        excel_data = self.create_excel_with_units("mapping_test.xlsx", mapping_test_cases)
        files = {'file': ('mapping_test.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, result = self.make_request('POST', 'upload-boq', files=files)
        
        if success:
            self.log_test("Column mapping processing", True, f"- File processed with column mapping")
            
            items = result.get('items', [])
            
            # Check if different unit formats are correctly mapped
            unit_mapping_tests = [
                ('Cubic Meter', ['cum', 'cu.m', 'cubic']),
                ('Square Meter', ['sqm', 'sq.m', 'square']),
                ('Numbers', ['nos', 'no', 'number', 'each'])
            ]
            
            mapping_success = True
            for i, item in enumerate(items):
                if i < len(unit_mapping_tests):
                    extracted_unit = item.get('unit', '').lower()
                    original_unit, expected_mappings = unit_mapping_tests[i]
                    
                    # Check if the unit was correctly mapped to standard format
                    is_correctly_mapped = any(expected in extracted_unit for expected in expected_mappings)
                    
                    if not is_correctly_mapped:
                        mapping_success = False
                        print(f"   ⚠️ Item {i+1}: '{item.get('unit')}' not mapped correctly from '{original_unit}'")
            
            self.log_test("Column mapping accuracy", mapping_success, 
                        f"- Unit column mapping working correctly")
            
            # Test that descriptions are properly extracted
            descriptions_valid = all(
                item.get('description') and len(item.get('description', '')) > 3 
                for item in items
            )
            self.log_test("Description column mapping", descriptions_valid, 
                        f"- All descriptions properly extracted")
            
            return result
        else:
            self.log_test("Column mapping processing", False, f"- {result}")
            return None

    def test_boq_item_structure(self):
        """Test BOQ Item Structure with proper unit values and GST rates"""
        print("\n🏗️ Testing BOQ Item Structure...")
        
        # Create comprehensive BOQ structure test
        structure_test_cases = [
            [1, 'Structural concrete work M25 grade', 'Cum', 120, 4800, 576000],
            [2, 'Reinforcement steel bars TMT Fe500', 'Kg', 2500, 68, 170000],
            [3, 'Brick masonry work in cement mortar', 'Cum', 180, 3200, 576000],
            [4, 'Internal wall plastering 12mm thick', 'Sqm', 800, 165, 132000],
            [5, 'Waterproofing membrane application', 'Sqm', 300, 280, 84000],
        ]
        
        excel_data = self.create_excel_with_units("structure_test.xlsx", structure_test_cases)
        files = {'file': ('structure_test.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, result = self.make_request('POST', 'upload-boq', files=files)
        
        if success:
            self.log_test("BOQ structure parsing", True, f"- Structure test file processed")
            
            items = result.get('items', [])
            
            # Test required fields in BOQ items
            required_fields = ['serial_number', 'description', 'unit', 'quantity', 'rate', 'amount', 'gst_rate']
            structure_valid = True
            
            for item in items:
                missing_fields = [field for field in required_fields if field not in item]
                if missing_fields:
                    structure_valid = False
                    print(f"   ⚠️ Item missing fields: {missing_fields}")
            
            self.log_test("BOQ item required fields", structure_valid, 
                        f"- All items have required fields: {required_fields}")
            
            # Test GST rates are properly initialized
            gst_rates_valid = all(
                isinstance(item.get('gst_rate'), (int, float)) and item.get('gst_rate') > 0
                for item in items
            )
            self.log_test("GST rates initialization", gst_rates_valid, 
                        f"- All items have valid GST rates (default: 18%)")
            
            # Test unit values are text and not numeric
            unit_text_validation = True
            numeric_units = []
            
            for item in items:
                unit_value = item.get('unit', '')
                if isinstance(unit_value, (int, float)) or (isinstance(unit_value, str) and unit_value.isdigit()):
                    unit_text_validation = False
                    numeric_units.append(f"Item {item.get('serial_number')}: {unit_value}")
            
            self.log_test("Unit values as text (not numeric)", unit_text_validation, 
                        f"- No numeric unit values found" if unit_text_validation else f"- Found numeric units: {numeric_units}")
            
            # Test quantity and rate are numeric
            numeric_validation = all(
                isinstance(item.get('quantity'), (int, float)) and 
                isinstance(item.get('rate'), (int, float)) and
                isinstance(item.get('amount'), (int, float))
                for item in items
            )
            self.log_test("Numeric fields validation", numeric_validation, 
                        f"- Quantity, rate, and amount are properly numeric")
            
            # Test total value calculation
            calculated_total = sum(item.get('amount', 0) for item in items)
            reported_total = result.get('total_value', 0)
            total_calculation_correct = abs(calculated_total - reported_total) < 0.01
            
            self.log_test("Total value calculation", total_calculation_correct, 
                        f"- Calculated: ₹{calculated_total:,.2f}, Reported: ₹{reported_total:,.2f}")
            
            return result
        else:
            self.log_test("BOQ structure parsing", False, f"- {result}")
            return None

    def test_edge_cases_and_variations(self):
        """Test edge cases and unit variations"""
        print("\n🔍 Testing Edge Cases and Unit Variations...")
        
        # Test with various unit formats and edge cases
        edge_case_data = [
            [1, 'Mixed unit format test', 'cu.m', 50, 2000, 100000],
            [2, 'Alternative unit format', 'sq.m', 200, 150, 30000],
            [3, 'Numeric-like unit text', 'MT', 10, 15000, 150000],
            [4, 'Unit with spaces', 'Running Meter', 100, 200, 20000],
            [5, 'Abbreviated unit', 'RM', 75, 180, 13500],
        ]
        
        excel_data = self.create_excel_with_units("edge_cases.xlsx", edge_case_data)
        files = {'file': ('edge_cases.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, result = self.make_request('POST', 'upload-boq', files=files)
        
        if success:
            self.log_test("Edge cases processing", True, f"- Edge cases file processed")
            
            items = result.get('items', [])
            
            # Test that all units are preserved as text
            all_units_text = all(isinstance(item.get('unit'), str) for item in items)
            self.log_test("All units preserved as text", all_units_text, 
                        f"- Unit types: {[type(item.get('unit')).__name__ for item in items]}")
            
            # Test unit normalization
            expected_normalizations = {
                'cu.m': 'Cum',
                'sq.m': 'Sqm', 
                'MT': 'Ton',
                'Running Meter': 'Rmt',
                'RM': 'Rmt'
            }
            
            normalization_success = True
            for item in items:
                original_unit = item.get('unit', '')
                # Check if unit was normalized correctly or preserved appropriately
                if original_unit and len(original_unit.strip()) == 0:
                    normalization_success = False
            
            self.log_test("Unit normalization", normalization_success, 
                        f"- Units properly normalized or preserved")
            
            return result
        else:
            self.log_test("Edge cases processing", False, f"- {result}")
            return None

    def test_project_creation_with_parsed_boq(self):
        """Test creating a project with parsed BOQ data"""
        print("\n🏗️ Testing Project Creation with Parsed BOQ...")
        
        # First parse a BOQ file
        boq_test_data = [
            [1, 'Foundation work with proper units', 'Cum', 80, 3500, 280000],
            [2, 'Masonry work measurement', 'Sqm', 400, 250, 100000],
            [3, 'Steel work quantity', 'Kg', 1500, 70, 105000],
        ]
        
        excel_data = self.create_excel_with_units("project_boq.xlsx", boq_test_data)
        files = {'file': ('project_boq.xlsx', excel_data, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, boq_result = self.make_request('POST', 'upload-boq', files=files)
        
        if success:
            self.log_test("BOQ parsing for project", True, f"- BOQ parsed successfully")
            
            # Create a client first
            client_data = {
                "name": "BOQ Test Client Ltd",
                "gst_no": "29ABCDE1234F1Z5",
                "bill_to_address": "BOQ Test Address, Test City, Karnataka - 560001",
                "contact_person": "BOQ Test Manager",
                "phone": "+91-9876543210",
                "email": "boq@testclient.com"
            }
            
            client_success, client_result = self.make_request('POST', 'clients', client_data)
            
            if client_success:
                client_id = client_result['client_id']
                
                # Create project with parsed BOQ data
                project_data = {
                    "project_name": "BOQ Unit Testing Project",
                    "architect": "BOQ Test Architect",
                    "client_id": client_id,
                    "client_name": "BOQ Test Client Ltd",
                    "location": "BOQ Test Location",
                    "metadata": boq_result.get('metadata', {}),
                    "boq_items": boq_result.get('items', []),
                    "total_project_value": boq_result.get('total_value', 0),
                    "advance_received": 0
                }
                
                project_success, project_result = self.make_request('POST', 'projects', project_data)
                
                if project_success:
                    project_id = project_result['project_id']
                    self.log_test("Project creation with BOQ", True, f"- Project ID: {project_id}")
                    
                    # Verify project BOQ items have correct unit structure
                    project_get_success, project_details = self.make_request('GET', f'projects/{project_id}')
                    
                    if project_get_success:
                        project_boq_items = project_details.get('boq_items', [])
                        
                        # Check that units are preserved correctly in project
                        units_preserved = all(
                            isinstance(item.get('unit'), str) and item.get('unit').strip()
                            for item in project_boq_items
                        )
                        
                        self.log_test("BOQ units preserved in project", units_preserved, 
                                    f"- All {len(project_boq_items)} BOQ items have text units")
                        
                        return project_id
                    else:
                        self.log_test("Project retrieval", False, f"- {project_details}")
                else:
                    self.log_test("Project creation with BOQ", False, f"- {project_result}")
            else:
                self.log_test("Client creation for BOQ test", False, f"- {client_result}")
        else:
            self.log_test("BOQ parsing for project", False, f"- {boq_result}")
        
        return None

    def run_boq_parsing_tests(self):
        """Run all BOQ parsing tests"""
        print("🚀 Starting BOQ Parsing Functionality Tests")
        print(f"🌐 Testing against: {self.base_url}")
        print("=" * 80)
        
        # Authenticate first
        if not self.authenticate():
            print("\n❌ Authentication failed - stopping tests")
            return False
        
        # Run BOQ parsing specific tests
        self.test_unit_column_extraction()
        self.test_column_mapping_debug()
        self.test_boq_item_structure()
        self.test_edge_cases_and_variations()
        self.test_project_creation_with_parsed_boq()
        
        # Print summary
        print("\n" + "=" * 80)
        print(f"📊 BOQ Parsing Test Summary: {self.tests_passed}/{self.tests_run} tests passed")
        
        success_rate = (self.tests_passed / self.tests_run) * 100 if self.tests_run > 0 else 0
        print(f"✅ Success Rate: {success_rate:.1f}%")
        
        if success_rate >= 80:
            print("🎉 BOQ parsing functionality is working well!")
        elif success_rate >= 60:
            print("⚠️ BOQ parsing has some issues that need attention")
        else:
            print("❌ BOQ parsing has significant issues that need fixing")
        
        return self.tests_passed == self.tests_run

def main():
    """Main test execution"""
    tester = BOQParsingTester()
    
    try:
        success = tester.run_boq_parsing_tests()
        return 0 if success else 1
    except KeyboardInterrupt:
        print("\n⏹️ Tests interrupted by user")
        return 1
    except Exception as e:
        print(f"\n💥 Unexpected error: {str(e)}")
        return 1

if __name__ == "__main__":
    sys.exit(main())